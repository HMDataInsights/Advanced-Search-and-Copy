# -*- coding: utf-8 -*-

import os
import shutil
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import openpyxl


def select_search_folder():
    global source_folder, L1
    source_folder = askdirectory()
    L1 = Label(root,text=source_folder, wraplength=300, justify = LEFT )
    L1.grid(row=5, column=1, sticky='w')
    return source_folder, L1


def count_files():
    global source_folder
    def count(source_folder):
        for entry in os.scandir(source_folder):
            if entry.is_dir(follow_symlinks=False):
                yield from count(entry.path)
            else:
                yield entry
    for i, j in enumerate(count(source_folder)):
        i = i
    return i


def select_dest_folder():
    global dest_folder, dest_file_dict, L2
    dest_file_dict = {}
    dest_folder = askdirectory()
    L2 = Label(root,text=dest_folder, wraplength=300, justify = LEFT)
    L2.grid(row=10, column=1, sticky='w')
    for name in os.listdir(dest_folder):
        path = os.path.join(dest_folder, name)
        if os.path.isfile(path):
            dest_file_dict[name] = 0
    print(dest_file_dict, '\n')
    return dest_folder, dest_file_dict, L2


def select_file_list():
    global sheet, L3
    file_list = askopenfilename(filetypes=[('Microsoft Excel Worksheet', '*.xlsx')])
    wb = openpyxl.load_workbook(file_list)
    L3 = Label(root, text=file_list, wraplength=300, justify=LEFT)
    L3.grid(row=15, column=1, sticky='w')
    sheet = wb.active
    return sheet, L3


def search_term():
    global file_name
    file_name = search_entry.get()
    file_name = file_name.replace('*', '')
    return file_name


def list_dir(folder_path):
    global file_path
    for name in os.listdir(folder_path):
        path = os.path.join(folder_path, name)
        if os.path.isdir(path):
            yield from list_dir(path)
        else:
            file_path = (os.path.join(folder_path, name))
            yield file_path


def list_files(folder_path):
    global file_path
    for entry in os.scandir(folder_path):
        if entry.is_dir():
            yield from list_files(entry)
        else:
            file_path = (entry.path)
            yield file_path


def search_file():
    global source_folder, dest_folder, sheet, file_name, dest_file_dict, tree
    progress['value'] = 0
    file_path = list_files(source_folder)
    try:
        file_name = search_term()
    except NameError: file_name = None
    try:
        file_list = sheet
    except NameError: file_list = None
    try:
        while True:
            name = next(file_path)
            if bool(file_name):
                if file_name in name:
                    copy_path = name
                    copied_file_name = os.path.basename(copy_path)
                    if variable_1.get() == 'Duplicate File Overwrite OFF':
                        if not os.path.exists(os.path.join(dest_folder, copied_file_name)):
                            shutil.copy(copy_path, dest_folder)
                            dest_file_dict[str(copied_file_name)] = 1
                        else:
                            dest_folder_1 = safe_copy(copied_file_name, dest_folder)
                            shutil.copy(copy_path, dest_folder_1[0])
                    else:
                        shutil.copy(copy_path, dest_folder)
                        dest_file_dict[str(copied_file_name)] = 1
            if bool(file_list):
                for i in range(1, sheet.max_row+2):
                    file_name = str(sheet.cell(row=i, column=1).value)
                    if file_name in name:
                        copy_path = name
                        copied_file_name = os.path.basename(copy_path)
                        if variable_1.get() == 'Duplicate File Overwrite OFF':
                            if not os.path.exists(os.path.join(dest_folder, copied_file_name)):
                                shutil.copy(copy_path, dest_folder)
                                dest_file_dict[str(copied_file_name)] = 1
                            else:
                                dest_folder_1 = safe_copy(copied_file_name, dest_folder)
                                shutil.copy(copy_path, dest_folder_1[0])
                        else:
                            shutil.copy(copy_path, dest_folder)
                            dest_file_dict[str(copied_file_name)] = 1
    except Exception:
        print(Exception())
        show_results()
        pass


def safe_copy(file_name, dest_folder):
    global dest_file_dict
    try:
        dest_file_dict[str(file_name)] = dest_file_dict.get(file_name)+1
    except TypeError:
        dest_file_dict[str(file_name)] = 1
    base, extension = os.path.splitext(file_name)
    dest_path = os.path.join(dest_folder, '{}_{}{}'.format(base, dest_file_dict.get(file_name), extension))
    return dest_path, dest_file_dict


def show_results():
    global tree
    try:
        tree.destroy()
    except NameError: pass
    i = 1
    global dest_file_dict
    tree = ttk.Treeview(root, selectmode='browse')
    tree.grid(column=0, columnspan=2)
    tree["columns"] = ("1", "2")
    tree.column("#0", minwidth=0, width=0, stretch='NO')
    tree.column("1", minwidth=300, anchor='center')
    tree.column("2", minwidth=75, anchor='center')
    tree.heading("#0", text="No.")
    tree.heading("1", text="File Name")
    tree.heading("2", text="No. of files found")
    vsb = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    vsb.place(x=410, y=170+150, height=50)

    # Inserting the items and their features to the columns built
    try:
        for item in dest_file_dict:
            if dest_file_dict[item] != 0:
                tree.insert("", 'end', text=i, values=(item, dest_file_dict[item]))
                i = i+1
    except NameError: pass


def bar():
    progress['value'] = 0
    for i in range(0, 125, 5):
        progress['value'] = i


def reset():
    global tree, source_folder, dest_folder, dest_file_dict, sheet, L1, L2, L3
    variable_1.set(OPTIONS_1[1])
    progress['value'] = 0
    try:
        L1.grid_forget()
    except NameError: pass
    try:
        L2.grid_forget()
    except NameError: pass
    try:
        L3.grid_forget()
    except NameError: pass
    try:
        del source_folder, dest_file_dict, sheet
    except NameError: pass
    tree.destroy()
    show_results()


def forget(widget):
    widget.forget()


def about():
    window = Toplevel(root)
    window.title("About Advanced Search and Copy V2.0")
    label_1 = Label(window, text=' Advanced Search and Copy is a simple tool to quickly search and copy an individual file or a batch of files to the specified folder.', wraplength=450, justify=CENTER, pady=10, font=("Helvetica", 10))
    label_1.pack()
    label_2 = Label(window, text=' \n Search File - Enter an individual file name to search', wraplength=450, justify=LEFT, pady=10, font=("Helvetica", 10))
    label_2.pack()
    label_3 = Label(window, text=' \n Search Folder - Select a folder to search files in. ', wraplength=450, justify=LEFT, pady=10, font=("Helvetica", 10))
    label_3.pack()
    label_4 = Label(window, text=' \n Destination Folder - Select a folder to copy the searched files in. ', wraplength=450, justify=LEFT, pady=10, font=("Helvetica", 10))
    label_4.pack()
    label_5 = Label(window, text=' \n Search file list - Select an excel file list to batch search. File list shall be in the first column.', wraplength=450, justify=LEFT, pady=10, font=("Helvetica", 10))
    label_5.pack()
    label_6 = Label(window, text=' \n Duplicate File Overwrite OFF - If a file with the same name already exists then the copied file name is altered to preserve both.' '\n Duplicate File Overwrite ON - If a file with the same name already exists in the destination folder, the copied file will overwrite the file in the destination folder.', wraplength=450, justify=LEFT, pady=10, font=("Helvetica", 10))
    label_6.pack()
    window.grab_set()


root = Tk()
root.geometry("430x375")
root.title("Advanced Search and Copy V1.0")
root.resizable(0, 0)
# root.configure(background='white')
''' Options menu'''
OPTIONS_1 = ['Duplicate File Overwrite ON', 'Duplicate File Overwrite OFF']
variable_1 = StringVar(root)
variable_1.set(OPTIONS_1[1])
w_1 = OptionMenu(root, variable_1, *OPTIONS_1)
w_1.grid(row=21, column=1, sticky='we')

'''Create Search Entry Box'''
search_entry =  Entry(root, width=50, bg='white')
search_entry.grid(row=0, column=1, sticky='we')
Label(root, text='Search File :', justify=LEFT).grid(row=0, column=0)

'''Menu Bar'''
menubar = Menu(root)
filemenu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="File", menu=filemenu)
filemenu.add_command(label="Exit", command=root.destroy)
helpmenu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="Help", menu=helpmenu)
helpmenu.add_command(label="About", command=about)
root.config(menu=menubar, )

'''Progress Bar'''
progress = ttk.Progressbar(root, orient='horizontal', length=120, value=0, mode='determinate')
progress.grid(row=22, column = 0, columnspan=2, sticky='we')
# progress.step(1)

button_1 = Button(root, text='Search Folder', borderwidth=1, command=select_search_folder)
button_1.grid(row=5, column=0, sticky='we')
button_1.flash()

button_2 = Button(root, text='Destination Folder', borderwidth=1, command=select_dest_folder)
button_2.grid(row=10, column=0, sticky='we')

button_3 = Button(root, text='Search Files List', borderwidth=1, command=select_file_list)
button_3.grid(row=15, column=0, sticky='we')


button_4 = Button(root, text='Search', borderwidth=1, command=lambda:[search_file(),bar()])
button_4.grid(row=20, column=0, sticky='we')

button_5 = Button(root, text='Reset', borderwidth=1, command=reset)
button_5.grid(row=21, column=0, sticky='we')

show_results()
root.mainloop()
